import csv
import io

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import (
    CreateView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

from lms_users.decorators import role_required
from lms_users.permissions import OwnerRequiredMixin, RoleRequiredMixin, Roles
from lms_users.services.keycloak import search_students

from ..forms import (
    CourseSemesterForm,
    EnrollmentForm,
    FinalAssignmentForm,
    LabParticipationGradeForm,
    LabSessionForm,
)
from ..models import (
    CourseSemester,
    FinalAssignment,
    FinalAssignmentResult,
    LabParticipation,
    LabReport,
    LabReportGrade,
    LabSession,
)


@method_decorator(login_required, name="dispatch")
class MyCourseSemestersList(RoleRequiredMixin, ListView):
    template_name = "lms_courses/teacher/my_course_semesters.html"
    context_object_name = "course_semesters"
    allowed_roles = (Roles.TEACHER,)

    def get_queryset(self):
        return CourseSemester.objects.select_related("course").filter(owner=self.request.user)


@method_decorator(login_required, name="dispatch")
class CourseSemesterCreate(RoleRequiredMixin, CreateView):
    template_name = "lms_courses/teacher/course_semester_create.html"
    form_class = CourseSemesterForm
    success_url = reverse_lazy("lms_courses_teacher:course_semester_list_teacher")
    allowed_roles = (Roles.TEACHER,)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request"] = self.request
        return kwargs


class CourseSemesterDetailView(OwnerRequiredMixin, TemplateView):
    template_name = "lms_courses/teacher/course_semester_detail.html"
    allowed_roles = (Roles.TEACHER,)
    model = CourseSemester

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cs = self.object  # type: ignore[attr-defined]
        # Prefetch sessions for all lab sessions (λιγότερα queries + σιγουριά ότι φαίνονται αμέσως)
        sessions = cs.sessions.all()  # type: ignore[attr-defined]
        ctx["course_semester"] = cs
        ctx["sessions"] = sessions
        ctx["students"] = cs.students.all()
        # Provide final_assignment (if implemented as a OneToOne with related_name).
        # Falls back to None so template shows the Create button.
        ctx["final_assignment"] = getattr(cs, "final_assignment", None)
        return ctx


class LabSessionManageView(RoleRequiredMixin, TemplateView):
    template_name = "lms_courses/teacher/lab_session_manage.html"
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.session = get_object_or_404(
            LabSession, pk=self.kwargs["session_id"], course_semester=self.cs
        )
        # Ensure report exists
        report = getattr(self.session, "report", None)
        if report is None:
            report = LabReport.objects.create(
                session=self.session,
                title=f"Report: {self.session.name}",
                max_grade=10,
                due_date=self.session.date,
            )
        self.report = report
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        students = list(self.cs.students.all().order_by("username"))
        # Build maps for current participation and grades
        parts = {
            p.student_id: p.present  # type: ignore
            for p in LabParticipation.objects.filter(session=self.session)
        }
        grades = {}
        if self.report:
            grades = {
                g.student_id: g.grade  # type: ignore
                for g in LabReportGrade.objects.filter(lab_report=self.report)
            }
        # Attach convenience attrs for template rendering (no leading underscores)
        for s in students:
            setattr(s, "present_value", bool(parts.get(s.id, False)))
            val = grades.get(s.id)
            setattr(s, "grade_value", "" if val is None else val)

        form = LabParticipationGradeForm(
            session=self.session,
            report=self.report,
            students_qs=self.cs.students.all(),
        )
        context = {
            "course_semester": self.cs,
            "session": self.session,
            "report": self.report,
            "students": students,
            "form": form,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = LabParticipationGradeForm(
            request.POST,
            session=self.session,
            report=self.report,
            students_qs=self.cs.students.all(),
        )
        if form.is_valid():
            form.save()

        return redirect(
            reverse("lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk})
        )


class LabSessionCreateView(RoleRequiredMixin, CreateView):
    template_name = "lms_courses/teacher/lab_session_form.html"
    form_class = LabSessionForm
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cy = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["course_semester"] = self.cy
        return kwargs

    def get_success_url(self):
        # Redirect to course semester detail view
        return reverse(
            "lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cy.pk}
        )


class EnrollmentCreateView(RoleRequiredMixin, FormView):
    template_name = "lms_courses/teacher/enrollment_form.html"
    form_class = EnrollmentForm
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["course_semester"] = self.cs
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["E2E_TEST_LOGIN"] = getattr(settings, "E2E_TEST_LOGIN", False)
        q = (self.request.GET.get("q") or "").strip()
        results = []
        if not ctx["E2E_TEST_LOGIN"] and q:
            try:
                results = search_students(q, max_results=10)
            except Exception:
                results = []
        ctx["query"] = q
        ctx["results"] = results
        return ctx

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse(
            "lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk}
        )


class FinalAssignmentCreateView(RoleRequiredMixin, CreateView):
    template_name = "lms_courses/teacher/final_assignment_form.html"
    form_class = FinalAssignmentForm
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        # Guard: if a final assignment already exists, redirect to edit
        existing = getattr(self.cs, "final_assignment", None)
        if existing is not None:
            return redirect(
                reverse("lms_courses_teacher:final_assignment_edit", kwargs={"pk": self.cs.pk})
            )
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["course_semester"] = self.cs
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Δημιουργία Τελικής Εργασίας"
        ctx["submit_label"] = "Δημιουργία"
        ctx["submit_testid"] = "submit-final-assignment"
        return ctx

    def get_success_url(self):
        return reverse(
            "lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk}
        )


class FinalAssignmentUpdateView(RoleRequiredMixin, UpdateView):
    template_name = "lms_courses/teacher/final_assignment_form.html"
    model = FinalAssignment
    form_class = FinalAssignmentForm
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.object = get_object_or_404(FinalAssignment, course_semester=self.cs)
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["instance"] = self.object
        kwargs["course_semester"] = self.cs
        return kwargs

    def get_object(self, queryset=None):
        # We already fetched the object in dispatch based on course_semester ownership
        return self.object

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Επεξεργασία Τελικής Εργασίας"
        ctx["submit_label"] = "Ενημέρωση"
        ctx["submit_testid"] = "submit-final-assignment"
        return ctx

    def get_success_url(self):
        return reverse(
            "lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk}
        )


class FinalAssignmentManageView(RoleRequiredMixin, TemplateView):
    template_name = "lms_courses/teacher/final_assignment_manage.html"
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.fa = get_object_or_404(FinalAssignment, course_semester=self.cs)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        students = list(self.cs.students.all().order_by("username"))
        # Current state maps
        res_map = {
            r.student_id: r  # type: ignore
            for r in FinalAssignmentResult.objects.filter(final_assignment=self.fa)
        }
        for s in students:
            r = res_map.get(s.id)
            setattr(s, "submitted_value", bool(getattr(r, "submitted", False)))
            setattr(s, "fa_grade_value", "" if r is None or r.grade is None else r.grade)

        context = {
            "course_semester": self.cs,
            "final_assignment": self.fa,
            "students": students,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        # Upsert results per student
        for student in self.cs.students.all():
            submitted = bool(request.POST.get(f"submitted_{student.id}") == "on")
            grade_val = request.POST.get(f"fa_grade_{student.id}")
            grade = int(grade_val) if grade_val not in (None, "") else None
            # Clamp grade within [0, max]
            if grade is not None:
                if grade < 0:
                    grade = 0
                if grade > self.fa.max_grade:
                    grade = self.fa.max_grade

            obj, _ = FinalAssignmentResult.objects.get_or_create(
                final_assignment=self.fa, student=student
            )
            changed = False
            if obj.submitted != submitted:
                obj.submitted = submitted
                changed = True
            if obj.grade != grade:
                obj.grade = grade
                changed = True
            if changed:
                obj.save()

        return redirect(
            reverse("lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk})
        )


class CourseSemesterDeleteView(RoleRequiredMixin, View):
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.object = get_object_or_404(
            CourseSemester.objects.all(), pk=self.kwargs["pk"], owner=request.user
        )
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object.delete()
        return redirect(reverse("lms_courses_teacher:course_semester_list_teacher"))


class LabSessionDeleteView(RoleRequiredMixin, View):
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.session = get_object_or_404(
            LabSession.objects.all(), pk=self.kwargs["session_id"], course_semester=self.cs
        )
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.session.delete()
        return redirect(
            reverse("lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk})
        )


class FinalAssignmentDeleteView(RoleRequiredMixin, View):
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.fa = getattr(self.cs, "final_assignment", None)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if self.fa:
            self.fa.delete()
        return redirect(
            reverse("lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk})
        )


class EnrollmentDeleteView(RoleRequiredMixin, View):
    allowed_roles = (Roles.TEACHER,)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        self.cs = get_object_or_404(
            CourseSemester.objects.select_related("course"),
            pk=self.kwargs["pk"],
            owner=request.user,
        )
        self.student_id = int(self.kwargs["student_id"])  # type: ignore[arg-type]
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Remove enrollment if present
        self.cs.students.remove(self.student_id)
        return redirect(
            reverse("lms_courses_teacher:course_semester_teacher_detail", kwargs={"pk": self.cs.pk})
        )


@role_required(Roles.TEACHER)
def export_course_semester(request, pk: int):
    cs = get_object_or_404(
        CourseSemester.objects.select_related("course"), pk=pk, owner=request.user
    )
    fmt = (request.GET.get("format") or "csv").lower()

    # Collect sessions, participations, grades, final assignment results
    sessions = list(cs.sessions.select_related("course_semester").all())  # type: ignore
    parts = list(
        LabParticipation.objects.filter(session__course_semester=cs)
        .select_related("student", "session")
        .order_by("session__week", "student__username")
    )
    grades = list(
        LabReportGrade.objects.filter(lab_report__session__course_semester=cs)
        .select_related("student", "lab_report__session")
        .order_by("lab_report__session__week", "student__username")
    )
    fa = getattr(cs, "final_assignment", None)
    fa_results = []
    if fa:
        fa_results = list(
            FinalAssignmentResult.objects.filter(final_assignment=fa)
            .select_related("student")
            .order_by("student__username")
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            fmt = "csv"
        else:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Στοιχεία"  # type: ignore
            ws1.append(["Πεδίο", "Τιμή"])  # type: ignore
            ws1.append(["Μάθημα", f"{cs.course.code} — {cs.course.title}"])  # type: ignore
            ws1.append(["Έτος", cs.year])  # type: ignore
            ws1.append(["Εξάμηνο", cs.get_semester_display()])  # type: ignore

            ws2 = wb.create_sheet("Συνεδρίες")
            ws2.append(["Εβδομάδα", "Όνομα", "Ημερομηνία", "Παρόντες", "Βαθμολογημένα"])
            for s in sessions:
                ws2.append(
                    [
                        s.week,
                        s.name,
                        getattr(s, "date", ""),
                        s.present_count,
                        s.graded_count,
                    ]
                )

            ws3 = wb.create_sheet("Παρουσίες")
            ws3.append(["Εβδομάδα", "Φοιτητής", "Παρουσία"])
            for p in parts:
                ws3.append([p.session.week, p.student.username, "Ναι" if p.present else "Όχι"])

            ws4 = wb.create_sheet("Βαθμοί εργαστηρίων")
            ws4.append(["Εβδομάδα", "Φοιτητής", "Βαθμός"])
            for g in grades:
                ws4.append([g.lab_report.session.week, g.student.username, g.grade or ""])

            ws5 = wb.create_sheet("Τελική εργασία")
            ws5.append(["Φοιτητής", "Υποβλήθηκε", "Βαθμός"])
            for r in fa_results:
                ws5.append([r.student.username, "Ναι" if r.submitted else "Όχι", r.grade or ""])

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(
                out.read(),
                content_type=(
                    "application/" "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            resp["Content-Disposition"] = f"attachment; filename=course_semester_{cs.pk}.xlsx"
            return resp

    # CSV
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["course_code", "course_title", "year", "semester"])
    writer.writerow([cs.course.code, cs.course.title, cs.year, cs.get_semester_display()])
    writer.writerow([])
    writer.writerow(["sessions: week", "name", "date", "present_count", "graded_count"])
    for s in sessions:
        writer.writerow([s.week, s.name, getattr(s, "date", ""), s.present_count, s.graded_count])
    writer.writerow([])
    writer.writerow(["participations: week", "student", "present"])
    for p in parts:
        writer.writerow([p.session.week, p.student.username, int(bool(p.present))])
    writer.writerow([])
    writer.writerow(["lab_grades: week", "student", "grade"])
    for g in grades:
        writer.writerow([g.lab_report.session.week, g.student.username, g.grade or ""])
    writer.writerow([])
    writer.writerow(["final_assignment: student", "submitted", "grade"])
    for r in fa_results:
        writer.writerow([r.student.username, int(bool(r.submitted)), r.grade or ""])

    csv_bytes = out.getvalue().encode("utf-8-sig")
    resp = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f"attachment; filename=course_semester_{cs.pk}.csv"
    return resp
