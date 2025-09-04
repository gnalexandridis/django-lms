from __future__ import annotations

import csv
import io

from django.core.cache import cache
from django.http import HttpResponse
from django.shortcuts import render

from lms.services.dashboard import compute_dashboard_stats
from lms_courses.models import CourseSemester
from lms_users.decorators import role_required
from lms_users.permissions import Roles


@role_required(Roles.TEACHER)
def teacher_home(request):
    user = request.user
    role = getattr(user, "role", None)
    ctx: dict = {"is_teacher": role in (Roles.TEACHER,)}

    # --- Filters ---
    try:
        days = int(request.GET.get("days", 7))
    except Exception:
        days = 7
    if days not in (3, 7, 14, 30):
        days = 7
    try:
        selected_course_id = int(request.GET.get("course", 0))
    except Exception:
        selected_course_id = 0

    # All courses for selector (for filter dropdown)
    all_courses_qs = CourseSemester.objects.filter(owner=user).select_related("course")

    cache_key = f"dash:{user.id}:d{days}:c{selected_course_id or 'all'}"
    cached = cache.get(cache_key)
    if cached:
        ctx.update({k: v for k, v in cached.items() if k not in ("all_courses_qs", "cs_qs")})
        ctx.update(
            {
                "filter_days": days,
                "filter_course_id": selected_course_id,
                "filter_days_options": [3, 7, 14, 30],
                "filter_courses": all_courses_qs,
            }
        )
        return render(request, "lms_courses/teacher/home.html", ctx)

    # Compute fresh stats via service
    computed = compute_dashboard_stats(user, days, selected_course_id)
    cache.set(cache_key, computed, 60)

    ctx.update({k: v for k, v in computed.items() if k not in ("all_courses_qs", "cs_qs")})
    ctx.update(
        {
            "filter_days": days,
            "filter_course_id": selected_course_id,
            "filter_days_options": [3, 7, 14, 30],
            "filter_courses": all_courses_qs,
        }
    )

    return render(request, "lms_courses/teacher/home.html", ctx)


def _export_dashboard(user, days: int, selected_course_id: int, fmt: str) -> HttpResponse:
    """Build dashboard export response for the given user/filters.

    This function is split out to make unit testing and coverage reliable
    independent of the @login_required/@role_required decorator chain.
    """
    data = compute_dashboard_stats(user, days, selected_course_id)

    per_course_rows = []
    for cs in data["per_course"]:
        per_course_rows.append(
            {
                "course_code": cs.course.code,
                "course_title": cs.course.title,
                "year": getattr(cs, "year", ""),
                "students": cs.students_count,
                "upcoming_sessions": cs.upcoming_sessions,
                "lab_done": cs.lab_done,
                "lab_null": cs.lab_null,
                "fa_sub": cs.fa_sub,
                "fa_grd": cs.fa_grd,
            }
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            fmt = "csv"
        else:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Στατιστικά"  # type: ignore
            ws1.append(["Κλειδί", "Τιμή"])  # type: ignore
            for key, val in [
                ("Ενεργά εξάμηνα", data["active_courses"]),
                ("Μοναδικοί φοιτητές", data["unique_students"]),
                ("Επερχόμενα labs", data["upcoming_labs"]),
                ("Lab ολοκληρωμένες", data["lab_grades_done"]),
                ("Lab εκκρεμείς", data["lab_grades_null"]),
                ("Final υποβολές", data["fa_submitted"]),
                ("Final βαθμολογημένες", data["fa_graded"]),
                ("Final μέσος βαθμός", data["fa_avg"] or ""),
                ("Καθυστερημένες", data["overdue_ungraded"]),
                ("Χωρίς παρουσίες", data["no_attendance_sessions"]),
            ]:
                ws1.append([key, val])  # type: ignore

            ws2 = wb.create_sheet("Ανά μάθημα")
            ws2.append(
                [
                    "Κωδικός",
                    "Τίτλος",
                    "Έτος",
                    "Φοιτητές",
                    "Επερχ.",
                    "Lab ✓",
                    "Lab εκκρ.",
                    "Final υποβ.",
                    "Final βαθμ.",
                ]
            )
            for r in per_course_rows:
                ws2.append(
                    [
                        r["course_code"],
                        r["course_title"],
                        r["year"],
                        r["students"],
                        r["upcoming_sessions"],
                        r["lab_done"],
                        r["lab_null"],
                        r["fa_sub"],
                        r["fa_grd"],
                    ]
                )

            ws3 = wb.create_sheet("Τάση")
            ws3.append(["Εβδομάδα-από", "Μετρήσεις"])
            for idx, v in enumerate(data["attendance_trend"], start=1):
                ws3.append([idx, v])

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(
                out.read(),
                content_type=(
                    "application/" "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            resp["Content-Disposition"] = (
                f"attachment; filename=dashboard_stats_d{days}_c{selected_course_id or 'all'}.xlsx"
            )
            return resp

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["key", "value"])
    writer.writerow(["active_courses", data["active_courses"]])
    writer.writerow(["unique_students", data["unique_students"]])
    writer.writerow(["upcoming_labs", data["upcoming_labs"]])
    writer.writerow(["lab_grades_done", data["lab_grades_done"]])
    writer.writerow(["lab_grades_null", data["lab_grades_null"]])
    writer.writerow(["fa_submitted", data["fa_submitted"]])
    writer.writerow(["fa_graded", data["fa_graded"]])
    writer.writerow(["fa_avg", data["fa_avg"] or ""])
    writer.writerow(["overdue_ungraded", data["overdue_ungraded"]])
    writer.writerow(["no_attendance_sessions", data["no_attendance_sessions"]])
    writer.writerow([])
    writer.writerow(
        [
            "course_code",
            "course_title",
            "year",
            "students",
            "upcoming_sessions",
            "lab_done",
            "lab_null",
            "fa_sub",
            "fa_grd",
        ]
    )
    for r in per_course_rows:
        writer.writerow(
            [
                r["course_code"],
                r["course_title"],
                r["year"],
                r["students"],
                r["upcoming_sessions"],
                r["lab_done"],
                r["lab_null"],
                r["fa_sub"],
                r["fa_grd"],
            ]
        )
    csv_bytes = out.getvalue().encode("utf-8-sig")
    resp = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = (
        f"attachment; filename=dashboard_stats_d{days}_c{selected_course_id or 'all'}.csv"
    )
    return resp


@role_required(Roles.TEACHER)
def teacher_export_dashboard(request):
    user = request.user
    role = getattr(user, "role", None)
    if role not in (Roles.TEACHER,):
        return HttpResponse(status=403)

    try:
        days = int(request.GET.get("days", 7))
    except Exception:
        days = 7
    if days not in (3, 7, 14, 30):
        days = 7
    try:
        selected_course_id = int(request.GET.get("course", 0))
    except Exception:
        selected_course_id = 0

    fmt = (request.GET.get("format") or "csv").lower()
    return _export_dashboard(user, days, selected_course_id, fmt)
